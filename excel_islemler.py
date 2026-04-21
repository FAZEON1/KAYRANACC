import pandas as pd
from io import BytesIO
from datetime import datetime


def excel_serial_to_date(n):
    """Excel seri numarasını tarihe çevirir."""
    try:
        if isinstance(n, (int, float)):
            from datetime import date, timedelta
            d = date(1899, 12, 30) + timedelta(days=int(n))
            return d.strftime("%Y-%m-%d")
        s = str(n).strip()[:10]
        pd.to_datetime(s)
        return s
    except Exception:
        return None


def parse_date(v):
    if v is None or v == "":
        return None
    # datetime/Timestamp objelerini direkt çevir
    try:
        if hasattr(v, 'strftime'):
            return v.strftime("%Y-%m-%d")
    except Exception:
        pass
    if isinstance(v, (int, float)):
        # Excel seri numarası olabilir (ama sadece makul aralıkta)
        if 10000 < float(v) < 80000:
            return excel_serial_to_date(v)
        return None
    try:
        s = str(v).strip()[:10]
        pd.to_datetime(s)
        return s
    except Exception:
        return None


def parse_num(v):
    """
    Sayı parse eder. Şu formatları destekler:
      - Gerçek int/float: 29298806.68
      - İngilizce format string: "29298806.68", "1,234.56"
      - Türkçe format string: "29.298.806,68", "1.234,56"
      - Para sembollü: "₺29.298.806,68", "$1,234.56"
    """
    import math

    if v is None or v == "":
        return None

    # numpy tipi ise native'e çevir
    if hasattr(v, 'item'):
        try:
            v = v.item()
        except Exception:
            pass

    # Zaten sayıysa direkt dön (Excel dtype=str değilse buraya düşer)
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return float(v)

    try:
        s = str(v).strip().replace(" ", "").replace("\xa0", "")
        if s.lower() in ("", "nan", "none", "-", "nat", "null"):
            return None

        # Para birimi sembollerini temizle
        for sym in ("₺", "$", "€", "£", "TL", "USD", "EUR", "try", "usd", "eur"):
            s = s.replace(sym, "")
        s = s.strip()
        if not s:
            return None

        has_comma = "," in s
        has_dot = "." in s

        if has_comma and has_dot:
            # Her iki ayraç var: son gelen ondalık ayraçtır
            if s.rfind(",") > s.rfind("."):
                # Türkçe: 1.234.567,89
                s = s.replace(".", "").replace(",", ".")
            else:
                # İngilizce: 1,234,567.89
                s = s.replace(",", "")
        elif has_comma:
            # Sadece virgül: Türkçe ondalık kabul et → "1234,56" → "1234.56"
            # (Binlik ayraç olarak kullanılmışsa zaten birden fazla virgül/nokta olurdu)
            s = s.replace(",", ".")
        elif has_dot:
            # Sadece nokta: birden fazla nokta varsa Türkçe binlik ayraç → sil
            # Tek nokta varsa ondalık ayraçtır, olduğu gibi bırak
            if s.count(".") > 1:
                s = s.replace(".", "")
            # Tek nokta varsa (örn. "29298806.68") float() zaten doğru okuyor

        result = float(s)
        if math.isnan(result) or math.isinf(result):
            return None
        return float(result)
    except Exception:
        return None


def normalize_kategori(k):
    """Kategori string'ini normalize eder."""
    if not k:
        return "diger"
    kategori_map = {
        "cek": "cek", "c": "cek",
        "kredi": "kredi",
        "kart": "kart", "k.karti": "kart", "kredi karti": "kart",
        "vergi": "vergi",
        "sgk": "sgk",
        "kira": "kira",
        "sabit": "sabit", "sabit gider": "sabit",
        "cari": "cari", "cari hesap": "cari",
        "ithalat": "ithalat",
        "ihracat": "ihracat",
        "masraf": "masraf",
        "maas": "maas", "odeme": "diger",
        "diger": "diger",
    }
    s = str(k).lower().strip()
    s = s.replace("ç", "c").replace("ş", "s").replace("ğ", "g")
    s = s.replace("ü", "u").replace("ö", "o").replace("ı", "i").replace("i̇", "i")
    return kategori_map.get(s, "diger")


def excel_yukle_odeme_listesi(file_bytes):
    """
    Haftalık ödeme listesi Excel'i okur.
    Sütun sırası: A=HAFTA, B=FİRMA, C=AÇIKLAMA, D=CARİ BANKA/IBAN, E=VADE, F=TUTAR TL, G=TUTAR USD, H=KATEGORİ

    ÖNEMLİ: dtype=str KULLANMAYIZ — değerlerin orijinal tipi (float, datetime) korunur.
    Bu sayede parse_num ondalık noktasını yanlış yorumlamaz.
    """
    try:
        # dtype=str YOK → sayılar float, tarihler datetime olarak gelir
        df = pd.read_excel(BytesIO(file_bytes), header=None)
        rows = df.values.tolist()

        hafta = ""
        if rows and rows[0] and rows[0][0] is not None:
            hv = rows[0][0]
            try:
                if hasattr(hv, 'strftime'):
                    hafta = hv.strftime("%d.%m.%Y")
                else:
                    hafta = str(hv).strip()
            except Exception:
                hafta = str(hv).strip()

        odemeler = []
        hatalar = []

        for i, r in enumerate(rows[2:], start=3):
            if not r or r[1] is None:
                continue
            firma_raw = r[1]
            # NaN / boş kontrolü
            try:
                if pd.isna(firma_raw):
                    continue
            except Exception:
                pass
            firma = str(firma_raw).strip()
            if not firma or firma.lower() in ("nan", "-", "none"):
                continue

            def safe_str(x):
                if x is None:
                    return ""
                try:
                    if pd.isna(x):
                        return ""
                except Exception:
                    pass
                return str(x).strip()

            aciklama = safe_str(r[2]) if len(r) > 2 else ""
            cari_banka = safe_str(r[3]) if len(r) > 3 else ""
            vade = parse_date(r[4]) if len(r) > 4 else None
            tl = parse_num(r[5]) if len(r) > 5 else None
            usd = parse_num(r[6]) if len(r) > 6 else None
            kategori_raw = r[7] if len(r) > 7 else None
            kategori = normalize_kategori(safe_str(kategori_raw)) if kategori_raw else "diger"

            if not tl and not usd:
                continue
            if not vade:
                hatalar.append(f"Satır {i}: '{firma}' için vade tarihi okunamadı.")
                continue

            odemeler.append({
                "firma": firma,
                "aciklama": aciklama,
                "cari_banka": cari_banka,
                "vade": vade,
                "tl": tl,
                "usd": usd,
                "kategori": kategori,
                "manuel": 0,
            })

        return hafta, odemeler, hatalar
    except Exception as e:
        return "", [], [f"Excel okuma hatası: {str(e)}"]


def excel_yukle_cek_listesi(file_bytes):
    """
    Sutun sirasi: A=Sira No, B=Referans No, C=Tarih, D=Vade Tarihi,
                  E=Cek No, F=Meblagh, G=Odenen, H=Kalan, I=Para Birimi,
                  J=Son Pozisyon, K=C/H Kodu, L=C/H Ismi, M=Banka, N=Sube, O=Hesap No
    """
    try:
        # Burada da dtype=str kullanmıyoruz
        df = pd.read_excel(BytesIO(file_bytes), header=None)
        rows = df.values.tolist()

        tl_cekler = []
        usd_cekler = []

        for i, r in enumerate(rows):
            if not r:
                continue
            # Tamamı boş satırı atla
            try:
                if all((v is None or (isinstance(v, float) and pd.isna(v))) for v in r):
                    continue
            except Exception:
                pass

            try:
                sira_raw = r[0]
                if sira_raw is None or (isinstance(sira_raw, float) and pd.isna(sira_raw)):
                    continue
                sira = float(str(sira_raw).strip())
                if sira <= 0:
                    continue
            except Exception:
                continue

            def cell(idx, default=""):
                if len(r) > idx and r[idx] is not None:
                    try:
                        if pd.isna(r[idx]):
                            return default
                    except Exception:
                        pass
                    v = r[idx]
                    if hasattr(v, 'strftime'):
                        return v.strftime("%Y-%m-%d")
                    s = str(v).strip()
                    if s.lower() in ("nan", "none"):
                        return default
                    return s
                return default

            ref_no = cell(1)
            if not ref_no:
                continue
            tarih = parse_date(r[2] if len(r) > 2 else None) or ""
            vade = parse_date(r[3] if len(r) > 3 else None) or ""
            cek_no = cell(4)
            meblagh = parse_num(r[5] if len(r) > 5 else None) or 0
            odenen = parse_num(r[6] if len(r) > 6 else None) or 0
            kalan = parse_num(r[7] if len(r) > 7 else None) or meblagh
            para_birimi = cell(8, "TL").upper().strip()
            pozisyon = cell(9, "Bekliyor")
            ch_kodu = cell(10)
            ch_ismi = cell(11)
            banka = cell(12)
            sube = cell(13)
            hesap_no = cell(14)

            if para_birimi not in ("TL", "USD"):
                para_birimi = "TL"

            cek = {
                "ref_no": ref_no,
                "cek_no": cek_no,
                "tarih": tarih,
                "vade": vade,
                "meblagh": meblagh,
                "odenen": odenen,
                "kalan": kalan,
                "para_birimi": para_birimi,
                "durum": pozisyon,
                "ch_kodu": ch_kodu,
                "ch_ismi": ch_ismi,
                "banka": banka,
                "sube": sube,
                "hesap_no": hesap_no,
            }
            if para_birimi == "USD":
                usd_cekler.append(cek)
            else:
                tl_cekler.append(cek)

        return tl_cekler, usd_cekler, []
    except Exception as e:
        return [], [], [f"Cek dosyasi okuma hatasi: {str(e)}"]


def export_excel(odemeler, hafta_adi, kur=38.5):
    """Ödeme listesini Excel'e aktarır."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Bu Hafta"

    # Başlık
    ws.merge_cells("A1:H1")
    ws["A1"] = hafta_adi or "Haftalık Ödeme Listesi"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0B1437")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Başlık satırı
    headers = ["FİRMA", "AÇIKLAMA", "KATEGORİ", "VADE", "TUTAR TL", "TUTAR USD", "DURUM", "ÖNCELİK"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="162050")
        cell.alignment = Alignment(horizontal="center")

    KATEGORI_LABELS = {
        "cek": "Çek", "kredi": "Kredi", "kart": "K.Kartı",
        "vergi": "Vergi", "sgk": "SGK", "kira": "Kira",
        "sabit": "Sabit Gider", "cari": "Cari Hesap", "diger": "Diğer"
    }
    DURUM_LABELS = {"odendi": "Ödendi", "bekliyor": "Bekliyor"}

    # Gün bazında grupla
    from collections import defaultdict
    by_day = defaultdict(list)
    for o in odemeler:
        day = (o.get("vade") or "")[:10] or "?"
        by_day[day].append(o)

    GUNLER = ["Pazar", "Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi"]

    row = 3
    tl_toplam = 0
    usd_toplam = 0

    for day in sorted(by_day.keys()):
        try:
            gun_adi = GUNLER[pd.to_datetime(day).weekday() + 1 if pd.to_datetime(day).weekday() < 6 else 0]
        except Exception:
            gun_adi = ""

        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = f"── {gun_adi} {day} ──"
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="1F2937")
        ws.row_dimensions[row].height = 20
        row += 1

        for o in by_day[day]:
            ws.cell(row=row, column=1, value=o.get("firma", ""))
            ws.cell(row=row, column=2, value=o.get("aciklama", ""))
            ws.cell(row=row, column=3, value=KATEGORI_LABELS.get(o.get("kategori", "diger"), "Diğer"))
            ws.cell(row=row, column=4, value=o.get("vade", ""))
            ws.cell(row=row, column=5, value=o.get("tutar_tl") or "")
            ws.cell(row=row, column=6, value=o.get("tutar_usd") or "")
            ws.cell(row=row, column=7, value=DURUM_LABELS.get(o.get("durum", "bekliyor"), "Bekliyor"))
            ws.cell(row=row, column=8, value=o.get("kategori", "diger"))
            if o.get("durum") == "odendi":
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="D1FAE5")
            tl_toplam += o.get("tutar_tl") or 0
            usd_toplam += o.get("tutar_usd") or 0
            row += 1

    # Toplam satırı
    row += 1
    ws.cell(row=row, column=1, value="TOPLAM")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=5, value=tl_toplam).font = Font(bold=True)
    ws.cell(row=row, column=6, value=usd_toplam).font = Font(bold=True)

    # Sütun genişlikleri
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def create_sample_excel():
    """Örnek Excel şablonu oluşturur."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Ödeme Listesi"

    ws["A1"] = "25. Hafta 16-22 Nisan 2026"
    ws.append([])
    ws.append(["HAFTA", "FİRMA", "AÇIKLAMA", "CARİ BANKA / IBAN", "VADE", "TUTAR TL", "TUTAR USD", "KATEGORİ"])

    for col in range(1, 9):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="162050")
        cell.alignment = Alignment(horizontal="center")

    ornekler = [
        ["", "ABC Lojistik", "Nakliye faturası", "TR12 0006 2001 1234 5678 9012 34", "2026-04-18", 45000, "", "cek"],
        ["", "XYZ Tedarik", "Ham madde", "TR98 0004 6004 0123 4567 8900 15", "2026-04-19", 120000, "", "cari"],
        ["", "Global Import", "USD odeme", "TR55 0001 0017 5432 1098 7654 32", "2026-04-21", "", 5000, "kredi"],
        ["", "Ofis Giderleri", "Kira Nisan", "TR33 0013 4000 9876 5432 1000 01", "2026-04-22", 28000, "", "kira"],
    ]
    for o in ornekler:
        ws.append(o)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

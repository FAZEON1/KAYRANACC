"""
rapor.py — KAYRANACC Raporlama Modülü
PDF ve Excel rapor üretimi
"""
from io import BytesIO
from datetime import datetime, date
import pandas as pd


KATEGORILER = {
    "cek":   {"label": "Çek",         "oncelik": 1},
    "kredi": {"label": "Kredi",        "oncelik": 2},
    "kart":  {"label": "K.Kartı",      "oncelik": 3},
    "vergi": {"label": "Vergi",        "oncelik": 4},
    "sgk":   {"label": "SGK",          "oncelik": 5},
    "kira":  {"label": "Kira",         "oncelik": 6},
    "sabit": {"label": "Sabit Gider",  "oncelik": 7},
    "cari":  {"label": "Cari Hesap",   "oncelik": 8},
    "diger": {"label": "Diğer",        "oncelik": 9},
}

GUNLER = ["Pazar", "Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi"]


def fmt(n):
    if n is None:
        return "-"
    try:
        return f"{float(n):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "-"


def fmt_tarih(s):
    if not s:
        return ""
    try:
        return pd.to_datetime(s).strftime("%d.%m.%Y")
    except Exception:
        return str(s)


def oncelik_sirala(o):
    kat = o.get("kategori") or "diger"
    return KATEGORILER.get(kat, {"oncelik": 9})["oncelik"]


# ── EXCEL RAPORU ──────────────────────────────────────────────────────

def haftalik_excel_raporu(odemeler, hafta_adi, bankalar=None, kur=38.5):
    """
    Tam haftalık Excel raporu üretir.
    Sayfa 1: Özet
    Sayfa 2: Günlük ödeme detayı
    Sayfa 3: Kategori analizi
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference

    wb = Workbook()

    # ── Renk sabitleri ──
    NAVY   = "0B1437"
    NAVY2  = "162050"
    GREEN  = "065F46"
    GREEN_BG = "D1FAE5"
    RED_BG   = "FEE2E2"
    AMBER_BG = "FEF3C7"
    BLUE_BG  = "DBEAFE"
    GRAY_BG  = "F3F4F6"

    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_style(cell, bg=NAVY):
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    def data_style(cell, bold=False, align="left", color=None):
        cell.font = Font(bold=bold, color=color or "111827", size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = border

    now_str = datetime.now().strftime("%d.%m.%Y %H:%M")

    # ═══════════════════════════════════════════════
    # SAYFA 1 — ÖZET
    # ═══════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Özet"
    ws1.sheet_view.showGridLines = False

    # Başlık
    ws1.merge_cells("A1:F1")
    ws1["A1"] = f"💳 KAYRANACC — {hafta_adi}"
    ws1["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws1["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:F2")
    ws1["A2"] = f"Oluşturulma: {now_str}"
    ws1["A2"].font = Font(size=10, color="6B7280")
    ws1["A2"].alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 18

    # Özet tablolar
    tl_toplam  = sum(o.get("tutar_tl")  or 0 for o in odemeler)
    usd_toplam = sum(o.get("tutar_usd") or 0 for o in odemeler)
    odendi_tl  = sum(o.get("tutar_tl")  or 0 for o in odemeler if o.get("durum") == "odendi")
    odendi_usd = sum(o.get("tutar_usd") or 0 for o in odemeler if o.get("durum") == "odendi")
    kalan_tl   = tl_toplam  - odendi_tl
    kalan_usd  = usd_toplam - odendi_usd
    odendi_cnt = sum(1 for o in odemeler if o.get("durum") == "odendi")

    ozet_rows = [
        ("Toplam TL Ödeme",  f"₺{fmt(tl_toplam)}",   None),
        ("Toplam USD Ödeme", f"${fmt(usd_toplam)}",   None),
        ("Ödendi TL",        f"₺{fmt(odendi_tl)}",    GREEN_BG),
        ("Ödendi USD",       f"${fmt(odendi_usd)}",   GREEN_BG),
        ("Bekleyen TL",      f"₺{fmt(kalan_tl)}",     AMBER_BG),
        ("Bekleyen USD",     f"${fmt(kalan_usd)}",     AMBER_BG),
        ("Ödeme Adedi",      f"{odendi_cnt}/{len(odemeler)}", None),
        ("Kur (USD/TL)",     f"{kur}",                None),
    ]

    ws1["A4"] = "Özet Bilgiler"
    ws1["A4"].font = Font(bold=True, size=12, color=NAVY)
    ws1.row_dimensions[4].height = 22

    for i, (label, val, bg) in enumerate(ozet_rows, start=5):
        ws1.cell(row=i, column=1, value=label)
        ws1.cell(row=i, column=2, value=val)
        data_style(ws1.cell(row=i, column=1), bold=True)
        c = ws1.cell(row=i, column=2)
        data_style(c, align="right")
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        ws1.row_dimensions[i].height = 18

    # Banka bakiyeleri
    if bankalar:
        row = 5 + len(ozet_rows) + 1
        ws1.cell(row=row, column=1, value="Banka Hesapları")
        ws1.cell(row=row, column=1).font = Font(bold=True, size=12, color=NAVY)
        ws1.row_dimensions[row].height = 22
        row += 1
        for b in bankalar:
            sym = "$" if b["para_birimi"] == "USD" else "₺"
            ws1.cell(row=row, column=1, value=b["hesap_adi"])
            ws1.cell(row=row, column=2, value=f"{sym}{fmt(b['bakiye'])} {b['para_birimi']}")
            data_style(ws1.cell(row=row, column=1), bold=True)
            data_style(ws1.cell(row=row, column=2), align="right")
            ws1.row_dimensions[row].height = 18
            row += 1

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 20

    # ═══════════════════════════════════════════════
    # SAYFA 2 — GÜNLÜK DETAY
    # ═══════════════════════════════════════════════
    ws2 = wb.create_sheet("Günlük Detay")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:H1")
    ws2["A1"] = f"{hafta_adi} — Günlük Ödeme Detayı"
    ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    headers = ["FİRMA", "AÇIKLAMA", "KATEGORİ", "VADE", "TUTAR TL (₺)", "TUTAR USD ($)", "DURUM", "AÇIKLAMA2"]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=2, column=col, value=h)
        header_style(c, NAVY2)
    ws2.row_dimensions[2].height = 20

    from collections import defaultdict
    by_day = defaultdict(list)
    for o in odemeler:
        day = (o.get("vade") or "")[:10] or "?"
        by_day[day].append(o)

    row = 3
    tl_top = 0
    usd_top = 0

    for day in sorted(by_day.keys()):
        try:
            d = pd.to_datetime(day)
            gun_adi = GUNLER[d.dayofweek + 1] if d.dayofweek < 6 else GUNLER[0]
            tarih_str = d.strftime("%d.%m.%Y")
        except Exception:
            gun_adi, tarih_str = "", day

        # Gün başlığı
        ws2.merge_cells(f"A{row}:H{row}")
        ws2[f"A{row}"] = f"── {gun_adi}  {tarih_str} ──"
        ws2[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws2[f"A{row}"].fill = PatternFill("solid", fgColor="1F2937")
        ws2[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws2.row_dimensions[row].height = 22
        row += 1

        for o in sorted(by_day[day], key=oncelik_sirala):
            is_odendi = o.get("durum") == "odendi"
            kat_label = KATEGORILER.get(o.get("kategori") or "diger", {"label": "Diğer"})["label"]
            fill_bg = GREEN_BG if is_odendi else None

            vals = [
                o.get("firma", ""),
                o.get("aciklama", ""),
                kat_label,
                fmt_tarih(o.get("vade")),
                o.get("tutar_tl"),
                o.get("tutar_usd"),
                "Ödendi ✓" if is_odendi else "Bekliyor",
                "",
            ]
            for col, val in enumerate(vals, 1):
                c = ws2.cell(row=row, column=col, value=val)
                data_style(c, align="right" if col in (5, 6) else "left",
                           bold=(col == 1))
                if fill_bg:
                    c.fill = PatternFill("solid", fgColor=fill_bg)

            tl_top  += o.get("tutar_tl")  or 0
            usd_top += o.get("tutar_usd") or 0
            ws2.row_dimensions[row].height = 18
            row += 1

    # Toplam satırı
    row += 1
    ws2.cell(row=row, column=1, value="GENEL TOPLAM").font = Font(bold=True, size=11)
    c_tl  = ws2.cell(row=row, column=5, value=tl_top)
    c_usd = ws2.cell(row=row, column=6, value=usd_top)
    for c in [c_tl, c_usd]:
        c.font = Font(bold=True, size=11)
        c.fill = PatternFill("solid", fgColor=GREEN_BG)
        c.alignment = Alignment(horizontal="right")

    col_widths = [35, 25, 14, 14, 18, 16, 12, 1]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════════════
    # SAYFA 3 — KATEGORİ ANALİZİ
    # ═══════════════════════════════════════════════
    ws3 = wb.create_sheet("Kategori Analizi")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:E1")
    ws3["A1"] = "Kategori Bazında Ödeme Analizi"
    ws3["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    kat_headers = ["KATEGORİ", "ÖDEME SAYISI", "TOPLAM TL (₺)", "TOPLAM USD ($)", "ÖDENEN ADET"]
    for col, h in enumerate(kat_headers, 1):
        c = ws3.cell(row=2, column=col, value=h)
        header_style(c, NAVY2)
    ws3.row_dimensions[2].height = 20

    kat_data = {}
    for o in odemeler:
        kat = o.get("kategori") or "diger"
        if kat not in kat_data:
            kat_data[kat] = {"cnt": 0, "tl": 0, "usd": 0, "odendi": 0}
        kat_data[kat]["cnt"] += 1
        kat_data[kat]["tl"]  += o.get("tutar_tl")  or 0
        kat_data[kat]["usd"] += o.get("tutar_usd") or 0
        if o.get("durum") == "odendi":
            kat_data[kat]["odendi"] += 1

    row = 3
    for kat, d in sorted(kat_data.items(), key=lambda x: KATEGORILER.get(x[0], {"oncelik": 9})["oncelik"]):
        label = KATEGORILER.get(kat, {"label": kat})["label"]
        cells = [label, d["cnt"], d["tl"] or None, d["usd"] or None, d["odendi"]]
        for col, val in enumerate(cells, 1):
            c = ws3.cell(row=row, column=col, value=val)
            data_style(c, align="right" if col > 1 else "left", bold=(col == 1))
        ws3.row_dimensions[row].height = 18
        row += 1

    for i, w in enumerate([20, 14, 18, 16, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── HTML / PDF RAPORU ─────────────────────────────────────────────────

def haftalik_html_raporu(odemeler, hafta_adi, bankalar=None, kur=38.5):
    """
    Tarayıcıda açılıp yazdırılabilir HTML rapor üretir.
    Streamlit'te st.download_button veya st.components ile kullanılır.
    """
    from collections import defaultdict

    now_str = datetime.now().strftime("%d.%m.%Y %H:%M")

    tl_toplam  = sum(o.get("tutar_tl")  or 0 for o in odemeler)
    usd_toplam = sum(o.get("tutar_usd") or 0 for o in odemeler)
    odendi_tl  = sum(o.get("tutar_tl")  or 0 for o in odemeler if o.get("durum") == "odendi")
    odendi_usd = sum(o.get("tutar_usd") or 0 for o in odemeler if o.get("durum") == "odendi")
    kalan_tl   = tl_toplam  - odendi_tl
    kalan_usd  = usd_toplam - odendi_usd
    odendi_cnt = sum(1 for o in odemeler if o.get("durum") == "odendi")

    # Banka özeti HTML
    banka_html = ""
    if bankalar:
        banka_html = '<div class="banka-grid">'
        for b in bankalar:
            sym = "$" if b["para_birimi"] == "USD" else "₺"
            banka_html += f"""
            <div class="banka-kart">
                <div class="banka-adi">{b['hesap_adi']}</div>
                <div class="banka-bakiye">{sym}{fmt(b['bakiye'])} <span class="banka-pb">{b['para_birimi']}</span></div>
            </div>"""
        banka_html += "</div>"

    # Ödeme satırları
    by_day = defaultdict(list)
    for o in odemeler:
        day = (o.get("vade") or "")[:10] or "?"
        by_day[day].append(o)

    tablo_rows = ""
    today = date.today().isoformat()

    KAT_RENKLER = {
        "cek":   "#dc2626", "kredi": "#ea580c", "kart":  "#d97706",
        "vergi": "#7c3aed", "sgk":   "#0891b2", "kira":  "#059669",
        "sabit": "#2563eb", "cari":  "#be185d", "diger": "#6b7280",
    }

    for day in sorted(by_day.keys()):
        try:
            d = pd.to_datetime(day)
            gun_adi  = GUNLER[d.dayofweek + 1] if d.dayofweek < 6 else GUNLER[0]
            tarih_str = d.strftime("%d.%m.%Y")
        except Exception:
            gun_adi, tarih_str = "", day

        is_today = day == today
        gun_row_bg = "#1e3a8a" if is_today else "#1a1a2e"

        tablo_rows += f"""
        <tr>
            <td colspan="5" style="background:{gun_row_bg};color:#fff;font-weight:700;
                                   padding:8px 12px;font-size:12px;letter-spacing:.3px">
                {'📅 BUGÜN — ' if is_today else ''}{gun_adi} — {tarih_str}
            </td>
        </tr>"""

        for o in sorted(by_day[day], key=oncelik_sirala):
            is_odendi = o.get("durum") == "odendi"
            kat = o.get("kategori") or "diger"
            kat_label = KATEGORILER.get(kat, {"label": "Diğer"})["label"]
            kat_renk  = KAT_RENKLER.get(kat, "#6b7280")
            row_style = "opacity:0.55;" if is_odendi else ""

            if o.get("tutar_tl"):
                tutar_html = f'<b style="color:#065F46">₺{fmt(o["tutar_tl"])}</b>'
            else:
                tutar_html = f'<b style="color:#1e40af">${fmt(o["tutar_usd"])}</b>'

            durum_bg    = "#dcfce7" if is_odendi else "#fef3c7"
            durum_color = "#166534" if is_odendi else "#92400e"
            durum_txt   = "✓ Ödendi" if is_odendi else "Bekliyor"

            tablo_rows += f"""
            <tr style="border-bottom:1px solid #e5e7eb;{row_style}">
                <td style="padding:8px 12px;border-left:4px solid {kat_renk}">
                    <div style="font-weight:700;font-size:12px">{o.get('firma','')}</div>
                    {'<div style="font-size:10px;color:#6b7280">' + o.get('aciklama','') + '</div>' if o.get('aciklama') else ''}
                </td>
                <td style="padding:8px 12px;text-align:center">
                    <span style="background:{kat_renk};color:#fff;font-size:10px;
                                 padding:2px 8px;border-radius:10px;font-weight:700">{kat_label}</span>
                </td>
                <td style="padding:8px 12px;font-size:11px;color:#6b7280;text-align:center">{fmt_tarih(o.get('vade'))}</td>
                <td style="padding:8px 12px;text-align:right">{tutar_html}</td>
                <td style="padding:8px 12px;text-align:center">
                    <span style="background:{durum_bg};color:{durum_color};
                                 font-size:10px;padding:2px 9px;border-radius:10px;font-weight:700">{durum_txt}</span>
                </td>
            </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="tr">
<head>
<meta charset="UTF-8">
<title>{hafta_adi} — KAYRANACC Rapor</title>
<style>
  * {{ box-sizing:border-box; margin:0; padding:0 }}
  body {{ font-family:'Segoe UI',Arial,sans-serif; background:#f4f5f7; color:#111; font-size:13px; padding:28px }}
  .header {{ background:linear-gradient(135deg,#0B1437,#162050); color:#fff; padding:20px 24px; border-radius:10px; margin-bottom:20px }}
  .header h1 {{ font-size:20px; margin:0 0 4px }}
  .header p  {{ font-size:11px; opacity:.6; margin:0 }}
  .ozet-grid {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(150px,1fr)); gap:12px; margin-bottom:20px }}
  .ozet-kart {{ background:#fff; border:1px solid #e5e7eb; border-radius:8px; padding:14px 16px; text-align:center }}
  .ozet-etiket {{ font-size:10px; color:#6b7280; text-transform:uppercase; letter-spacing:.5px; margin-bottom:6px }}
  .ozet-deger  {{ font-size:18px; font-weight:700 }}
  .banka-grid {{ display:flex; gap:12px; flex-wrap:wrap; margin-bottom:20px }}
  .banka-kart {{ background:#fff; border:1px solid #e5e7eb; border-radius:8px; padding:12px 16px; min-width:160px }}
  .banka-adi  {{ font-size:10px; color:#9ca3af; text-transform:uppercase; margin-bottom:4px }}
  .banka-bakiye {{ font-size:18px; font-weight:700; font-family:monospace }}
  .banka-pb   {{ font-size:11px; color:#9ca3af }}
  table {{ width:100%; border-collapse:collapse; background:#fff; border-radius:8px; overflow:hidden; box-shadow:0 1px 4px rgba(0,0,0,.07) }}
  th {{ background:#f3f4f6; padding:9px 12px; text-align:left; font-size:11px; color:#6b7280; border-bottom:2px solid #e5e7eb; font-weight:700; text-transform:uppercase; letter-spacing:.3px }}
  .section-title {{ font-size:13px; font-weight:700; color:#0B1437; text-transform:uppercase; letter-spacing:.5px; margin:18px 0 10px }}
  @media print {{
    body {{ padding:12px; background:white }}
    .header {{ border-radius:0; -webkit-print-color-adjust:exact; print-color-adjust:exact }}
  }}
</style>
</head>
<body>

<div class="header">
  <h1>💳 KAYRANACC — {hafta_adi}</h1>
  <p>Rapor tarihi: {now_str} · USD/TL Kur: {kur}</p>
</div>

<div class="ozet-grid">
  <div class="ozet-kart">
    <div class="ozet-etiket">Toplam TL</div>
    <div class="ozet-deger" style="color:#065F46">₺{fmt(tl_toplam)}</div>
  </div>
  <div class="ozet-kart">
    <div class="ozet-etiket">Toplam USD</div>
    <div class="ozet-deger" style="color:#1e40af">${fmt(usd_toplam)}</div>
  </div>
  <div class="ozet-kart">
    <div class="ozet-etiket">Ödendi TL</div>
    <div class="ozet-deger" style="color:#059669">₺{fmt(odendi_tl)}</div>
  </div>
  <div class="ozet-kart">
    <div class="ozet-etiket">Kalan TL</div>
    <div class="ozet-deger" style="color:#d97706">₺{fmt(kalan_tl)}</div>
  </div>
  <div class="ozet-kart">
    <div class="ozet-etiket">Kalan USD</div>
    <div class="ozet-deger" style="color:#d97706">${fmt(kalan_usd)}</div>
  </div>
  <div class="ozet-kart">
    <div class="ozet-etiket">İlerleme</div>
    <div class="ozet-deger">{odendi_cnt}/{len(odemeler)}</div>
  </div>
</div>

{f'<div class="section-title">🏦 Banka Bakiyeleri</div>{banka_html}' if banka_html else ''}

<div class="section-title">📋 Ödeme Detayları</div>
<table>
  <tr>
    <th>Firma</th>
    <th>Kategori</th>
    <th style="text-align:center">Vade</th>
    <th style="text-align:right">Tutar</th>
    <th style="text-align:center">Durum</th>
  </tr>
  {tablo_rows}
</table>

<div style="text-align:center;margin-top:24px;font-size:10px;color:#9ca3af">
  KAYRANACC · {now_str}
</div>

<script>
  // Otomatik yazdır (isteğe bağlı kaldırılabilir)
  // window.onload = () => setTimeout(() => window.print(), 500);
</script>
</body>
</html>"""

    return html.encode("utf-8")


def nakit_akis_excel(odemeler, bankalar, hafta_adi, kur=38.5):
    """Nakit akış tablosunu Excel'e aktarır."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    wb = Workbook()
    ws = wb.active
    ws.title = "Nakit Akış"
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Nakit Akış — {hafta_adi}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0B1437")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Tarih", "Günlük TL (₺)", "Günlük USD ($)", "Kümülatif TL (₺)", "Kümülatif USD ($)", "Bakiye Kalan (₺)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="162050")
        c.alignment = Alignment(horizontal="center")
        c.border = border
    ws.row_dimensions[2].height = 20

    banka_tl = sum(b["bakiye"] for b in bankalar if b["para_birimi"] == "TL")

    by_day = defaultdict(list)
    for o in odemeler:
        if o.get("durum") == "bekliyor":
            day = (o.get("vade") or "")[:10] or "?"
            by_day[day].append(o)

    kum_tl = kum_usd = 0
    row = 3
    for day in sorted(by_day.keys()):
        gun_tl  = sum(o.get("tutar_tl")  or 0 for o in by_day[day])
        gun_usd = sum(o.get("tutar_usd") or 0 for o in by_day[day])
        kum_tl  += gun_tl
        kum_usd += gun_usd
        kalan = banka_tl - kum_tl - (kum_usd * kur)

        vals = [day, gun_tl or None, gun_usd or None, kum_tl, kum_usd, kalan]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = border
            c.alignment = Alignment(horizontal="right" if col > 1 else "left")
            c.font = Font(size=10)
            if col == 6 and kalan < 0:
                c.fill = PatternFill("solid", fgColor="FFCCCC")
                c.font = Font(bold=True, color="C62828")
        ws.row_dimensions[row].height = 18
        row += 1

    net = banka_tl - kum_tl - (kum_usd * kur)
    ws.cell(row=row, column=1, value="TOPLAM").font = Font(bold=True)
    ws.cell(row=row, column=4, value=kum_tl).font = Font(bold=True)
    ws.cell(row=row, column=5, value=kum_usd).font = Font(bold=True)
    c = ws.cell(row=row, column=6, value=net)
    c.font = Font(bold=True, color="065F46" if net >= 0 else "C62828")
    c.fill = PatternFill("solid", fgColor="D1FAE5" if net >= 0 else "FFCCCC")

    for i, w in enumerate([14, 18, 16, 18, 16, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
